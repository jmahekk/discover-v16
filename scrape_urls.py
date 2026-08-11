"""Scrapes paper titles and links from the ACL Anthology into hyperlinked Excel files."""

import argparse
import os
import time
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font

CONFERENCES = {
    "acl": {
        "url":     "https://aclanthology.org/events/acl-2025/",
        "section": "2025acl-long",
        "out":     "input/acl_2025_urls.xlsx",
    },
    "eacl": {
        "url":     "https://aclanthology.org/events/eacl-2026/",
        "section": "2026eacl-long",
        "out":     "input/eacl_2026_urls.xlsx",
    },
    "emnlp": {
        "url":     "https://aclanthology.org/events/emnlp-2025/",
        "section": "2025emnlp-main",
        "out":     "input/emnlp_2025_urls.xlsx",
    },
    "naacl": {
        "url":     "https://aclanthology.org/events/naacl-2025/",
        "section": "2025naacl-long",
        "out":     "input/naacl_2025_urls.xlsx",
    },
    "conll": {
        "url":     "https://aclanthology.org/events/conll-2025/",
        "section": "2025conll-1",
        "out":     "input/conll_2025_urls.xlsx",
    },
    "findings": {
        "url":     "https://aclanthology.org/events/findings-2026/",
        "section": "2026findings-eacl",
        "out":     "input/findings_2026_urls.xlsx",
    },
}

HEADERS = {"User-Agent": "Mozilla/5.0"}


def fetch_papers(page_url: str, section_id: str) -> list:
    """
    Downloads the anthology event page, finds the correct section,
    and returns a list of {"title": ..., "url": ...} for every paper in it.
    """
    print(f"  Fetching: {page_url}")
    r = requests.get(page_url, headers=HEADERS, timeout=30)
    r.raise_for_status()

    soup = BeautifulSoup(r.text, "html.parser")

    # Confirm the section exists on the page
    section_start = soup.find(id=section_id)
    if section_start is None:
        raise ValueError(
            f"Section '{section_id}' not found on page. "
            f"Open {page_url} in your browser and verify the section id."
        )

    # Build the paper URL prefix from section_id.
    # Examples:
    #   "2025acl-long"      ->  "/2025.acl-long"
    #   "2025emnlp-main"    ->  "/2025.emnlp-main"
    #   "2025conll-1"       ->  "/2025.conll-1"
    #   "2026findings-eacl" ->  "/2026.findings-eacl"
    year   = section_id[:4]        # "2025"
    rest   = section_id[4:]        # "acl-long"
    prefix = f"/{year}.{rest}"     # "/2025.acl-long"

    papers = []
    seen   = set()

    for a in soup.find_all("a", href=True):
        href = a["href"]

        if not href.startswith(prefix):
            continue

        tail = href[len(prefix):]
        if not tail or tail == "/":
            continue              # skip the volume-level link, not a paper

        if href in seen:
            continue

        title = a.get_text(strip=True)
        if not title:
            continue

        # Skip bib download links only
        if title.lower() == "bib":
            continue
        if href.endswith(".bib"):
            continue

        seen.add(href)
        full_url = "https://aclanthology.org" + href.rstrip("/")
        papers.append({"title": title, "url": full_url})

    return papers


def save_to_excel(papers: list, out_path: str):
    """
    Saves papers to xlsx where column A = title with a hyperlink.
    This is exactly the format that read_urls.py expects.
    """
    os.makedirs(os.path.dirname(out_path), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Papers"

    ws["A1"] = "Title"
    ws["A1"].font = Font(bold=True)

    link_font = Font(color="0000FF", underline="single")

    for i, paper in enumerate(papers, start=2):
        cell           = ws.cell(row=i, column=1, value=paper["title"])
        cell.hyperlink = paper["url"]
        cell.font      = link_font

    wb.save(out_path)
    print(f"  Saved {len(papers)} papers -> {out_path}")


def scrape_one(key: str):
    cfg = CONFERENCES[key]
    print(f"\n[{key.upper()}]")
    try:
        papers = fetch_papers(cfg["url"], cfg["section"])
        if not papers:
            print(f"  WARNING: 0 papers found. Check section id '{cfg['section']}'.")
            return
        save_to_excel(papers, cfg["out"])
        time.sleep(1)
    except Exception as e:
        print(f"  ERROR: {e}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--conf",
        choices=list(CONFERENCES.keys()),
        default=None,
        help="Scrape only one conference. Omit to scrape all 6.",
    )
    args = parser.parse_args()

    os.makedirs("input", exist_ok=True)

    if args.conf:
        scrape_one(args.conf)
    else:
        for key in CONFERENCES:
            scrape_one(key)

    print("\nDone. Check the input/ folder for the xlsx files.")


if __name__ == "__main__":
    main()
