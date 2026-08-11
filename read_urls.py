"""Reads the hyperlinked title and URL pairs out of a scraped Excel file."""

from openpyxl import load_workbook

def read_hyperlinks(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    urls = []

    for row in ws.iter_rows(min_row=1):
        cell = row[0]   
        if cell.hyperlink:
            urls.append({
                "title": cell.value,
                "url": cell.hyperlink.target
            })

    return urls
